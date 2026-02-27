"""
Sitemark — PDF Pre-commissioning → Excel avec photos
Interface graphique, packageable en .exe via PyInstaller
(CI build: push de ce fichier déclenche Build Sitemark.exe)
"""
import sys
import os
import re
import io
import threading
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

import fitz
import pdfplumber
import requests
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

IMG_W, IMG_H = 220, 165
STATUS_COLORS  = {"Résolu": "C6EFCE", "En cours": "FFEB9C", "À faire": "FFC7CE"}
STATUS_DARK    = {"Résolu": "2E7D32", "En cours": "E65100", "À faire": "C62828"}
GRAVITY_COLORS = {"1": "FF5252", "2": "FF9800", "3": "FFC107"}

KNOWN_LABELS = [
    ("Date de création", "date_creation"),
    ("Description de la remarque", "description"),
    ("Localisation de la réserve", "localisation"),
    ("Photo 1 de la réserve", "photo1_r"), ("Photo 2 de la réserve", "photo2_r"), ("Photo 3 de la réserve", "photo3_r"),
    ("Photo 1 de levée constaté", "photo1_l"), ("Photo 2 de levée constaté", "photo2_l"), ("Photo 3 de levée constaté", "photo3_l"),
    ("Responsable concerné", "responsable"), ("ID du composant", "id_composant"),
    ("Mis à jour à", "date_maj"), ("Sélectionner", "type_reserve"),
    ("Attribué à", "attribue_a"), ("Créé Par", "cree_par"),
    ("Longitude", "longitude"), ("Latitude", "latitude"),
    ("Gravité", "gravite"), ("Rangée", "rangee"), ("Statut", "statut"),
    ("Tables", "tables"), ("Zone", "zone"), ("ID", "id_reserve"),
]

def match_label(text):
    t = text.strip()
    for label, key in KNOWN_LABELS:
        if t == label or t.startswith(label + " "):
            return key, t[len(label):].strip()
    return None, None

def words_text(words):
    return " ".join(w["text"] for w in sorted(words, key=lambda w: w["x0"]))

def seg(row_words, x_min, x_max=9999):
    return [w for w in row_words if x_min <= w["x0"] < x_max]

def extract_all(pdf_path):
    reserves, page_map, photo_urls_map = [], {}, {}
    id_re = re.compile(r"^#(\d+)$")
    doc = fitz.open(pdf_path)

    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            m = re.search(r"^#(\d+)$", text, re.MULTILINE)
            if m:
                rid = int(m.group(1))
                page_map[rid] = {"data_page": pi, "photo_page": pi + 1}

    for pi in range(len(doc)):
        links = doc[pi].get_links()
        urls = [l["uri"] for l in links if "sitemark.com" in l.get("uri", "")]
        if urls:
            for rid, info in page_map.items():
                if info.get("photo_page") == pi:
                    photo_urls_map[rid] = urls
                    break

    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue
            rows = {}
            for w in words:
                y = round(w["top"] / 3) * 3
                rows.setdefault(y, []).append(w)
            y_keys = sorted(rows.keys())
            pending_key = None

            for idx, y_key in enumerate(y_keys):
                row_words = sorted(rows[y_key], key=lambda w: w["x0"])
                row_texts = [w["text"] for w in row_words]

                if len(row_words) == 1 and id_re.match(row_texts[0]):
                    reserves.append({"_id": int(id_re.match(row_texts[0]).group(1))})
                    pending_key = None
                    continue

                if not reserves:
                    continue
                cur = reserves[-1]
                left_words = seg(row_words, 0, 290)
                right_words = seg(row_words, 290)

                if words_text(right_words).strip() == "Sélectionner" and not left_words:
                    prev = words_text(seg(sorted(rows[y_keys[idx-1]], key=lambda w: w["x0"]), 390)) if idx > 0 else ""
                    nxt  = words_text(seg(sorted(rows[y_keys[idx+1]], key=lambda w: w["x0"]), 390)) if idx+1 < len(y_keys) else ""
                    cur["type_reserve"] = (prev + " " + nxt).strip()
                    pending_key = None
                    continue

                if words_text(row_words).strip() in ("manquant", "manquante") and not left_words:
                    continue

                if left_words:
                    lt = words_text(left_words)
                    key, rem = match_label(lt)
                    if key:
                        cur[key] = rem.strip()
                        pending_key = key if not rem.strip() else None
                    elif pending_key:
                        cur[pending_key] = (cur.get(pending_key, "") + " " + lt).strip()
                        pending_key = None

                if right_words:
                    rt = words_text(right_words)
                    if rt.strip() != "Sélectionner":
                        key, rem = match_label(rt)
                        if key:
                            cur[key] = rem.strip()

    for r in reserves:
        ir = r.get("id_reserve", "")
        if not ir or not re.match(r"^\d+$", ir.strip()):
            r["id_reserve"] = str(r.get("_id", ""))
        for k in ["photo1_r", "photo2_r", "photo3_r", "photo1_l", "photo2_l", "photo3_l"]:
            r.pop(k, None)
        r.pop("_id", None)
        for k in list(r.keys()):
            if isinstance(r[k], str):
                r[k] = r[k].strip()

    return reserves, photo_urls_map

def download_image(url):
    try:
        resp = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        img = PILImage.open(io.BytesIO(resp.content))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        img.thumbnail((IMG_W, IMG_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        return buf
    except:
        return None

def bdr():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def convert(pdf_path, out_path, on_progress, on_done, on_error):
    try:
        on_progress(5, "Lecture du PDF...")
        reserves, photo_urls_map = extract_all(pdf_path)
        site = os.path.basename(pdf_path).replace(".pdf", "")
        now  = datetime.today().strftime("%d/%m/%Y %H:%M")
        total = sum(len(v) for v in photo_urls_map.values())

        on_progress(15, f"{len(reserves)} réserves — téléchargement de {total} photos...")

        photo_images = {}
        done = 0
        for rid, urls in photo_urls_map.items():
            photo_images[rid] = []
            for url in urls:
                buf = download_image(url)
                photo_images[rid].append(buf)
                done += 1
                pct = 15 + int(done / max(total, 1) * 55)
                on_progress(pct, f"Photos : {done}/{total}")

        on_progress(72, "Génération de l'Excel...")
        wb = Workbook()
        b  = bdr()

        ws1 = wb.active
        ws1.title = "Réserves"
        META = [
            ("ID", "id_reserve", 6), ("Statut", "statut", 14), ("Gravité", "gravite", 9),
            ("Type Réserve", "type_reserve", 34), ("Description", "description", 32),
            ("Localisation", "localisation", 22), ("Zone", "zone", 8), ("Rangée", "rangee", 9),
            ("Tables", "tables", 8), ("Créé Par", "cree_par", 18), ("Attribué À", "attribue_a", 18),
            ("Responsable", "responsable", 16), ("Date Création", "date_creation", 16),
            ("Date MAJ", "date_maj", 16), ("Nb Photos", "_nb", 10),
        ]
        nc = len(META)
        lc = get_column_letter(nc)

        def mh(ws, row, text, bg="1F3864", fg="FFFFFF", sz=12, h=28):
            ws.merge_cells(f"A{row}:{lc}{row}")
            c = ws[f"A{row}"]
            c.value = text
            c.font  = Font(name="Calibri", bold=True, size=sz, color=fg)
            c.fill  = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = h

        mh(ws1, 1, f"Pré-commissioning — {site}")
        mh(ws1, 2, f"Généré le {now}  |  {len(reserves)} réserves", bg="D9E1F2", fg="595959", sz=9, h=18)

        for ci, (hdr, _, w) in enumerate(META, 1):
            c = ws1.cell(row=3, column=ci, value=hdr)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", start_color="2E75B6")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = b
            ws1.column_dimensions[get_column_letter(ci)].width = w
        ws1.row_dimensions[3].height = 32

        for ri, r in enumerate(reserves, 4):
            rid = int(r.get("id_reserve", 0) or 0)
            alt = PatternFill("solid", start_color="EBF3FB" if ri % 2 == 0 else "FFFFFF")
            for ci, (_, fk, _) in enumerate(META, 1):
                v = str(len(photo_images.get(rid, []))) if fk == "_nb" else str(r.get(fk, "") or "")
                cell = ws1.cell(row=ri, column=ci, value=v)
                cell.font = Font(name="Calibri", size=9)
                cell.border = b
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if fk == "statut" and v in STATUS_COLORS:
                    cell.fill = PatternFill("solid", start_color=STATUS_COLORS[v])
                    cell.font = Font(name="Calibri", size=9, bold=True)
                elif fk == "gravite" and v in GRAVITY_COLORS:
                    cell.fill = PatternFill("solid", start_color=GRAVITY_COLORS[v])
                    cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
                else:
                    cell.fill = alt
            ws1.row_dimensions[ri].height = 22
        ws1.freeze_panes = "A4"

        on_progress(80, "Insertion des photos...")

        ws2 = wb.create_sheet("Photos")
        ws2.merge_cells("A1:H1")
        c = ws2["A1"]
        c.value = f"Photos — {site}"
        c.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28
        ws2.column_dimensions["A"].width = 18
        for ci in range(2, 10):
            ws2.column_dimensions[get_column_letter(ci)].width = int(IMG_W / 7) + 2

        cur_row = 2
        for r in reserves:
            rid    = int(r.get("id_reserve", 0) or 0)
            statut = r.get("statut", "")
            type_r = r.get("type_reserve", "")
            imgs   = photo_images.get(rid, [])

            ws2.merge_cells(f"A{cur_row}:H{cur_row}")
            hc = ws2[f"A{cur_row}"]
            hc.value = f"#{rid}  |  {statut}  |  {type_r}  |  {len(imgs)} photo(s)"
            hc.font  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            hc.fill  = PatternFill("solid", start_color=STATUS_DARK.get(statut, "37474F"))
            hc.alignment = Alignment(vertical="center", indent=1)
            ws2.row_dimensions[cur_row].height = 22
            cur_row += 1

            if not imgs:
                ws2.cell(row=cur_row, column=1, value="Aucune photo").font = Font(italic=True, size=9, color="999999")
                ws2.row_dimensions[cur_row].height = 18
                cur_row += 2
                continue

            ws2.row_dimensions[cur_row].height = IMG_H + 8
            lbl = ws2.cell(row=cur_row, column=1, value="Photos →")
            lbl.font = Font(name="Calibri", size=8, italic=True, color="555555")
            lbl.alignment = Alignment(vertical="top")

            for i, buf in enumerate(imgs):
                if buf is None:
                    continue
                xl = XLImage(buf)
                xl.width  = IMG_W
                xl.height = IMG_H
                ws2.add_image(xl, f"{get_column_letter(i + 2)}{cur_row}")
            cur_row += 2

        ws3 = wb.create_sheet("Logs")
        ws3.merge_cells("A1:F1")
        c = ws3["A1"]
        c.value = f"Logs — {site} — {now}"
        c.font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color="37474F")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 24
        hdrs = [("Réserve ID", 12), ("Statut", 14), ("Gravité", 10), ("Type", 34), ("Description", 40), ("Nb Photos", 10)]
        for ci, (h, w) in enumerate(hdrs, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", start_color="37474F")
            c.border = b
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[2].height = 22
        for ri, r in enumerate(reserves, 3):
            rid  = int(r.get("id_reserve", 0) or 0)
            vals = [r.get("id_reserve",""), r.get("statut",""), r.get("gravite",""),
                    r.get("type_reserve",""), r.get("description",""), str(len(photo_images.get(rid,[])))]
            alt  = PatternFill("solid", start_color="F5F5F5" if ri % 2 == 0 else "FFFFFF")
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=v)
                cell.font = Font(name="Calibri", size=9)
                cell.fill = alt
                cell.border = b
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws3.row_dimensions[ri].height = 18
        ws3.freeze_panes = "A3"

        ws4 = wb.create_sheet("Synthèse")
        statuts, gravites, types = {}, {}, {}
        for r in reserves:
            for d, k in [(statuts,"statut"), (gravites,"gravite"), (types,"type_reserve")]:
                v = str(r.get(k,"?") or "?"); d[v] = d.get(v, 0) + 1

        def wt(ws, row, title, data):
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, name="Calibri")
            ws.cell(row=row+1, column=1, value="Valeur").font = Font(bold=True, name="Calibri")
            ws.cell(row=row+1, column=2, value="Nb").font = Font(bold=True, name="Calibri")
            for i, (k, v) in enumerate(sorted(data.items()), row+2):
                ws.cell(row=i, column=1, value=k)
                ws.cell(row=i, column=2, value=v)
            return row + len(data) + 4

        r = wt(ws4, 1, "Par Statut", statuts)
        r = wt(ws4, r, "Par Gravité", gravites)
        wt(ws4, r, "Par Type", types)
        ws4.column_dimensions["A"].width = 40
        ws4.column_dimensions["B"].width = 8
        ws4.cell(row=1, column=4, value="Total photos").font = Font(bold=True, name="Calibri")
        ws4.cell(row=2, column=4, value=total)

        on_progress(95, "Sauvegarde...")
        wb.save(out_path)
        on_progress(100, "Terminé !")
        on_done(out_path)

    except Exception as e:
        import traceback
        on_error(str(e) + "\n" + traceback.format_exc())


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sitemark — Convertisseur PDF")
        self.geometry("520x360")
        self.resizable(False, False)
        self.configure(bg="#1F3864")

        tk.Label(self, text="Sitemark", font=("Calibri", 24, "bold"),
                 bg="#1F3864", fg="white").pack(pady=(30, 2))
        tk.Label(self, text="PDF Pré-commissioning  →  Excel avec photos",
                 font=("Calibri", 11), bg="#1F3864", fg="#A8C4E0").pack()

        frame = tk.Frame(self, bg="white")
        frame.pack(fill="both", expand=True, padx=30, pady=20)

        self.pdf_var = tk.StringVar(value="Aucun fichier sélectionné")
        tk.Label(frame, textvariable=self.pdf_var, font=("Calibri", 9),
                 bg="white", fg="#666666", wraplength=420).pack(pady=(20, 10))

        self.btn = tk.Button(
            frame,
            text="  Choisir un PDF Sitemark  →  Convertir",
            font=("Calibri", 12, "bold"),
            bg="#2E75B6", fg="white",
            activebackground="#1F5496", activeforeground="white",
            relief="flat", cursor="hand2", padx=20, pady=14,
            command=self.run)
        self.btn.pack(pady=6)

        style = ttk.Style()
        style.theme_use("default")
        style.configure("blue.Horizontal.TProgressbar",
                         thickness=12, troughcolor="#EEEEEE", background="#2E75B6")
        self.progress = ttk.Progressbar(frame, length=420, mode="determinate",
                                         style="blue.Horizontal.TProgressbar")
        self.progress.pack(pady=(18, 4))

        self.status_var = tk.StringVar(value="Prêt")
        tk.Label(frame, textvariable=self.status_var, font=("Calibri", 9, "italic"),
                 bg="white", fg="#888888").pack()

    def run(self):
        pdf_path = filedialog.askopenfilename(
            title="Choisir le rapport Sitemark PDF",
            filetypes=[("Fichiers PDF", "*.pdf")])
        if not pdf_path:
            return

        self.pdf_var.set(os.path.basename(pdf_path))
        out_path = os.path.join(
            os.path.dirname(pdf_path),
            os.path.splitext(os.path.basename(pdf_path))[0] + "_reserves.xlsx")

        self.btn.config(state="disabled")
        self.progress["value"] = 0

        def on_progress(pct, msg):
            self.progress["value"] = pct
            self.status_var.set(msg)
            self.update_idletasks()

        def on_done(path):
            self.btn.config(state="normal")
            self.status_var.set("✓ Terminé !")
            if messagebox.askyesno("Succès",
                    f"Conversion terminée !\n\nFichier :\n{path}\n\nOuvrir ?"):
                os.startfile(path)

        def on_error(msg):
            self.btn.config(state="normal")
            self.status_var.set("Erreur")
            messagebox.showerror("Erreur", f"Erreur :\n\n{msg[:500]}")

        threading.Thread(
            target=convert,
            args=(pdf_path, out_path, on_progress, on_done, on_error),
            daemon=True).start()


if __name__ == "__main__":
    App().mainloop()
